import pandas as pd
import webbrowser
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class FileHandler:
    
    def write_to_text(self, q_table, output_path):

        try:
            with open(output_path, 'w', encoding='utf-8') as file:
                for i, v in q_table.items():
                    file.write(f"{i}: {v}\n")

            print(f"Veriler '{output_path}' dosyasına başarıyla yazıldı.")

        except IOError as e:
            print(f"Dosya yazılırken bir hata oluştu: {e}")



    def write_qvalue_updates_to_excel(self, all_episodes_info, file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "qvalue_updates_2col"

        # --- Stiller (Değişiklik yok) ---
        font_bold = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        thick_side = Side(style='thick', color='000000')
        light_gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        dark_gray_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        dim_gray_fill = PatternFill(start_color="696969", end_color="696969", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        start_col = 1
        num_columns = 2  # Her bölüm için sütun sayısı 2'ye indirildi

        for episode_info in all_episodes_info:
            episode_col = start_col

            # --- Başlıklar (YENİ DÜZEN) ---
            # Bölüm başlığını 2 sütuna birleştir
            ws.merge_cells(start_row=1, start_column=episode_col, end_row=1, end_column=episode_col + num_columns - 1)
            title_cell = ws.cell(row=1, column=episode_col)
            title_cell.value = f"Episode {episode_info['episode_num'] + 1}"
            title_cell.font = font_bold
            title_cell.alignment = center_alignment
            title_cell.fill = light_gray_fill

            # Diğer bilgileri 2 sütun kullanarak alt alta yaz
            header_info = {
                "Experienced": episode_info['switch_point'],
                "Model-Selected": episode_info['model_selected_switching_point'],
                "Explored": episode_info['explored_switching_point'] if episode_info['explored_switching_point'] is not None else 'None',
                "Result": episode_info.get('termination_type', 'Normal'),
                "Total Time": episode_info['total_time'],
                "Total Weight": episode_info['final_weight']
            }
            
            current_row = 2
            for label, value in header_info.items():
                # Etiket hücresi
                label_cell = ws.cell(row=current_row, column=episode_col)
                label_cell.value = f"{label}:"
                label_cell.font = font_bold
                
                # Değer hücresi
                value_cell = ws.cell(row=current_row, column=episode_col + 1)
                value_cell.value = value
                value_cell.font = font_bold
                
                # "Result" satırını renklendir
                if label == "Result":
                    termination_fill = {
                        "Normal": green_fill, "Underflow": yellow_fill, "Overflow": red_fill
                    }.get(value, green_fill)
                    label_cell.fill = termination_fill
                    value_cell.fill = termination_fill

                current_row += 1

            # Sütun başlıkları (Weight, Q Value)
            data_header_row = current_row
            columns = ["Weight", "Q Value"]
            fills = [silver_fill, dim_gray_fill]
            for idx, (col_name, fill) in enumerate(zip(columns, fills)):
                cell = ws.cell(row=data_header_row, column=episode_col + idx)
                cell.value = col_name
                cell.font = font_bold
                cell.fill = fill
                cell.alignment = center_alignment

            # --- Q-Değeri Verilerini Yazma (Güncellenmiş) ---
            row_idx = data_header_row + 1
            experienced_weight = episode_info['switch_point']
            
            # Sıralamayı anahtara (key) göre yap (önceki cevaptaki gibi)
            # episode_info['q_value'] şeklinde düzeltildiğini varsayıyorum
            sorted_items = sorted(episode_info['q_value'].items(), key=lambda item: item[0])
            
            for weight, q_val in sorted_items:
                # action ve count artık yazılmıyor
                ws.cell(row=row_idx, column=episode_col, value=weight).font = font_bold
                
                q_value_cell = ws.cell(row=row_idx, column=episode_col + 1)
                q_value_cell.value = float(f"{q_val:.4f}")
                q_value_cell.font = font_bold

                if weight == experienced_weight:
               
                    for offset in range(num_columns):
                        ws.cell(row=row_idx, column=episode_col + offset).fill = highlight_fill
                else:
                 
                    ws.cell(row=row_idx, column=episode_col).fill = light_gray_fill
                    ws.cell(row=row_idx, column=episode_col + 1).fill = dark_gray_fill
                
                row_idx += 1

            # --- Formatlama (Güncellenmiş) ---
            # Sütun genişlikleri
            for idx in range(num_columns):
                col_letter = get_column_letter(episode_col + idx)
                ws.column_dimensions[col_letter].width = 20

            # Kenarlıklar
            top_row = 1
            bottom_row = row_idx - 1
            for row in range(top_row, bottom_row + 1):
                for col_num in range(num_columns):
                    col = episode_col + col_num
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if row == top_row:
                        cell.border = Border(top=thick_side, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
                    if row == bottom_row:
                        cell.border = Border(bottom=thick_side, left=cell.border.left, right=cell.border.right, top=cell.border.top)
                    if col_num == 0:
                        cell.border = Border(left=thick_side, top=cell.border.top, bottom=cell.border.bottom, right=cell.border.right)
                    if col_num == num_columns - 1:
                        cell.border = Border(right=thick_side, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

            # Sonraki blok için başlangıç sütununu ayarla
            start_col += num_columns

        # Yeni başlık düzenine göre dondurulmuş bölmeyi ayarla
        ws.freeze_panes = f"A{data_header_row + 1}"
        wb.save(file_path)